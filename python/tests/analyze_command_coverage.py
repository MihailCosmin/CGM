#!/usr/bin/env python3
"""
CGM Command Coverage Analyzer

Analyzes all cleartext CGM files in batch_tests folder and generates an Excel report
showing which commands are handled in the Python SVG converter and what they convert to.

Usage:
    python analyze_command_coverage.py

Output:
    command_coverage_report.xlsx - Excel file with command analysis
"""

import os
import re
import sys
from pathlib import Path
from collections import defaultdict, Counter
from typing import Dict, List, Set, Tuple
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class CommandAnalyzer:
    """Analyzes CGM commands and their SVG conversions"""
    
    # Define command mappings based on cleartextcgm_to_svg.py
    COMMAND_MAPPINGS = {
        # Structural commands
        'BEGMF': 'XML Comment (Metafile)',
        'BEGPIC': 'XML Comment (Picture)',
        'BEGPICBODY': 'No output (structure)',
        'ENDPIC': 'No output (structure)',
        'ENDMF': 'No output (structure)',
        'BEGAPS': 'No output (structure)',
        'BEGAPSBODY': 'No output (structure)',
        'ENDAPS': 'No output (structure)',
        'BEGFIG': 'State: in_figure=True',
        'BEGFIGURE': 'State: in_figure=True',
        'ENDFIG': 'State: in_figure=False',
        'ENDFIGURE': 'State: in_figure=False',
        
        # Coordinate system
        'vdcext': 'State: vdc_extent (viewport)',
        'MAXVDCEXT': 'State: max_vdc_extent',
        'scalemode': 'State: metric_scale_factor',
        
        # Graphical primitives
        'LINE': 'SVG: <line> or <polyline>',
        'DISJTLINE': 'SVG: multiple <line>',
        'POLYGON': 'SVG: <polygon>',
        'CIRCLE': 'SVG: <circle>',
        'ELLIPSE': 'SVG: <ellipse>',
        'ARCCTR': 'SVG: <path> (arc)',
        'ELLIPARC': 'SVG: <path> (elliptical arc)',
        'POLYBEZIER': 'SVG: <path> (bezier curves)',
        
        # Text
        'TEXT': 'SVG: <text> with transform',
        'RESTRTEXT': 'SVG: <text> with transform',
        
        # Line attributes
        'linewidth': 'State: line_width',
        'linetype': 'State: line_type',
        'linecolr': 'State: line_color',
        'LINECAP': 'State: line_cap',
        'LINEJOIN': 'State: line_join',
        'linewidthmode': 'No output (setup)',
        
        # Edge attributes
        'edgevis': 'State: edge_visible',
        'EDGEWIDTH': 'State: edge_width',
        'EDGECOLR': 'State: edge_color',
        'EDGETYPE': 'No output (setup)',
        'EDGEVIS': 'State: edge_visible',
        'EDGEWIDTHMODE': 'No output (setup)',
        
        # Fill attributes
        'fillcolr': 'State: fill_color',
        'intstyle': 'State: interior_style',
        
        # Text attributes
        'textcolr': 'State: text_color',
        'textfontindex': 'State: text_font_index',
        'charheight': 'State: character_height',
        'charori': 'State: character_orientation',
        'CHARORI': 'State: character_orientation',
        'textalign': 'State: text_alignment',
        'TEXTALIGN': 'State: text_alignment',
        
        # Color
        'COLRTABLE': 'State: color_table',
        'colrtable': 'State: color_table',
        'backcolr': 'No output (setup)',
        
        # Clipping
        'CLIP': 'State: clip_rectangle',
        
        # Setup/metadata commands (no rendering output)
        'MFVERSION': 'No output (setup)',
        'MFDESC': 'No output (setup)',
        'MFELEMLIST': 'No output (setup)',
        'fontlist': 'No output (setup)',
        'CHARSETLIST': 'No output (setup)',
        'VDCTYPE': 'No output (setup)',
        'COLRPREC': 'No output (setup)',
        'COLRINDEXPREC': 'No output (setup)',
        'COLRVALUEEXT': 'No output (setup)',
        'MAXCOLRINDEX': 'No output (setup)',
        'INTEGERPREC': 'No output (setup)',
        'REALPREC': 'No output (setup)',
        'charcoding': 'No output (setup)',
        'colrmode': 'No output (setup)',
        'VDCREALPREC': 'No output (setup)',
        'ALTCHARSETINDEX': 'No output (setup)',
        'CHARSETINDEX': 'No output (setup)',
        'HATCHSTYLEDEF': 'No output (setup)',
        'PATTERNDEFN': 'No output (setup)',
        'INTERPINT': 'No output (setup)',
        'TRANSPARENCY': 'No output (setup)',
        'APSATTR': 'No output (setup)',
        'RESTRTEXTTYPE': 'No output (setup)',
        'MESSAGE': 'No output (ignored)',
    }
    
    def __init__(self, batch_tests_dir: Path):
        self.batch_tests_dir = batch_tests_dir
        self.file_commands: Dict[str, List[str]] = {}  # filename -> list of commands
        self.all_commands: Set[str] = set()  # all unique commands found
        self.command_counts: Counter = Counter()  # command -> count across all files
        
    def extract_command_name(self, line: str) -> str:
        """Extract command name from CGM line"""
        line = line.strip()
        if not line or line.startswith('%'):
            return None
        
        # Get the first word/token
        match = re.match(r'^([A-Za-z0-9]+)', line)
        if match:
            return match.group(1)
        return None
    
    def analyze_file(self, cgm_file: Path) -> List[str]:
        """Analyze a single cleartext CGM file and return list of commands"""
        commands = []
        
        try:
            with open(cgm_file, 'r', encoding='utf-8', errors='ignore') as f:
                for line in f:
                    cmd = self.extract_command_name(line)
                    if cmd:
                        commands.append(cmd)
                        self.all_commands.add(cmd)
                        self.command_counts[cmd] += 1
        except Exception as e:
            print(f"Error reading {cgm_file.name}: {e}")
            
        return commands
    
    def analyze_all_files(self):
        """Analyze all cleartext CGM files in batch_tests directory"""
        cleartext_files = sorted(self.batch_tests_dir.glob("*_cleartext.cgm"))
        
        print(f"Found {len(cleartext_files)} cleartext CGM files")
        print("Analyzing files...")
        
        for i, cgm_file in enumerate(cleartext_files, 1):
            print(f"  [{i}/{len(cleartext_files)}] {cgm_file.name}")
            commands = self.analyze_file(cgm_file)
            self.file_commands[cgm_file.name] = commands
        
        print(f"\nFound {len(self.all_commands)} unique commands")
        print(f"Total command instances: {sum(self.command_counts.values())}")
    
    def get_command_mapping(self, command: str) -> str:
        """Get SVG conversion mapping for a command"""
        # Check exact match first
        if command in self.COMMAND_MAPPINGS:
            return self.COMMAND_MAPPINGS[command]
        
        # Check case-insensitive match
        for key, value in self.COMMAND_MAPPINGS.items():
            if key.upper() == command.upper():
                return value
        
        # Not handled
        return "NOT HANDLED"
    
    def generate_excel_report(self, output_file: Path):
        """Generate Excel report with command coverage analysis"""
        print(f"\nGenerating Excel report: {output_file}")
        
        # Create workbook
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create sheets
        self._create_summary_sheet(wb)
        self._create_command_list_sheet(wb)
        self._create_file_by_file_sheet(wb)
        self._create_unhandled_commands_sheet(wb)
        
        # Save workbook
        wb.save(output_file)
        print(f"✓ Report saved: {output_file}")
    
    def _create_summary_sheet(self, wb):
        """Create summary overview sheet"""
        ws = wb.create_sheet("Summary", 0)
        
        # Title
        ws['A1'] = "CGM Command Coverage Analysis"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:D1')
        
        # Statistics
        row = 3
        total_commands = len(self.all_commands)
        handled_commands = sum(1 for cmd in self.all_commands 
                              if self.get_command_mapping(cmd) != "NOT HANDLED")
        unhandled_commands = total_commands - handled_commands
        
        stats = [
            ("Total Files Analyzed:", len(self.file_commands)),
            ("Total Unique Commands:", total_commands),
            ("Handled Commands:", handled_commands),
            ("Unhandled Commands:", unhandled_commands),
            ("Coverage Percentage:", f"{(handled_commands/total_commands*100):.1f}%" if total_commands > 0 else "N/A"),
            ("Total Command Instances:", sum(self.command_counts.values())),
        ]
        
        for label, value in stats:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Top 10 most common commands
        row += 2
        ws[f'A{row}'] = "Top 10 Most Common Commands"
        ws[f'A{row}'].font = Font(size=12, bold=True)
        row += 1
        
        ws[f'A{row}'] = "Command"
        ws[f'B{row}'] = "Count"
        ws[f'C{row}'] = "SVG Conversion"
        for cell in [ws[f'A{row}'], ws[f'B{row}'], ws[f'C{row}']]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        row += 1
        
        for cmd, count in self.command_counts.most_common(10):
            ws[f'A{row}'] = cmd
            ws[f'B{row}'] = count
            ws[f'C{row}'] = self.get_command_mapping(cmd)
            
            # Color code
            if self.get_command_mapping(cmd) == "NOT HANDLED":
                ws[f'C{row}'].fill = PatternFill(start_color="FFB3B3", end_color="FFB3B3", fill_type="solid")
            
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 40
    
    def _create_command_list_sheet(self, wb):
        """Create sheet with all commands and their mappings"""
        ws = wb.create_sheet("Command List")
        
        # Headers
        headers = ["Command", "Total Count", "SVG Conversion", "Status"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Data
        row = 2
        for cmd in sorted(self.all_commands):
            mapping = self.get_command_mapping(cmd)
            status = "✓ Handled" if mapping != "NOT HANDLED" else "✗ Not Handled"
            
            ws[f'A{row}'] = cmd
            ws[f'B{row}'] = self.command_counts[cmd]
            ws[f'C{row}'] = mapping
            ws[f'D{row}'] = status
            
            # Color code status
            if status == "✗ Not Handled":
                ws[f'D{row}'].fill = PatternFill(start_color="FFB3B3", end_color="FFB3B3", fill_type="solid")
                ws[f'D{row}'].font = Font(color="CC0000", bold=True)
            else:
                ws[f'D{row}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                ws[f'D{row}'].font = Font(color="006100")
            
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 45
        ws.column_dimensions['D'].width = 15
    
    def _create_file_by_file_sheet(self, wb):
        """Create sheet with command breakdown per file"""
        ws = wb.create_sheet("File-by-File Analysis")
        
        # Headers
        ws['A1'] = "File Name"
        ws['B1'] = "Command"
        ws['C1'] = "Count in File"
        ws['D1'] = "SVG Conversion"
        
        for cell in [ws['A1'], ws['B1'], ws['C1'], ws['D1']]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Data
        row = 2
        for filename in sorted(self.file_commands.keys()):
            commands = self.file_commands[filename]
            file_cmd_counts = Counter(commands)
            
            # Start row for this file
            start_row = row
            
            for cmd in sorted(file_cmd_counts.keys()):
                count = file_cmd_counts[cmd]
                mapping = self.get_command_mapping(cmd)
                
                ws[f'A{row}'] = filename if row == start_row else ""
                ws[f'B{row}'] = cmd
                ws[f'C{row}'] = count
                ws[f'D{row}'] = mapping
                
                # Color code
                if mapping == "NOT HANDLED":
                    ws[f'D{row}'].fill = PatternFill(start_color="FFB3B3", end_color="FFB3B3", fill_type="solid")
                
                row += 1
            
            # Add border after each file
            if row > start_row:
                for col in range(1, 5):
                    ws.cell(row=row-1, column=col).border = Border(
                        bottom=Side(style='thick', color='000000')
                    )
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 45
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 45
    
    def _create_unhandled_commands_sheet(self, wb):
        """Create sheet highlighting unhandled commands"""
        ws = wb.create_sheet("Unhandled Commands")
        
        # Find unhandled commands
        unhandled = sorted([cmd for cmd in self.all_commands 
                           if self.get_command_mapping(cmd) == "NOT HANDLED"])
        
        # Title
        ws['A1'] = f"Unhandled Commands ({len(unhandled)} total)"
        ws['A1'].font = Font(size=14, bold=True, color="CC0000")
        ws.merge_cells('A1:C1')
        
        if not unhandled:
            ws['A3'] = "✓ All commands are handled!"
            ws['A3'].font = Font(size=12, bold=True, color="006100")
            ws['A3'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            return
        
        # Headers
        row = 3
        ws[f'A{row}'] = "Command"
        ws[f'B{row}'] = "Total Count"
        ws[f'C{row}'] = "Files Containing This Command"
        
        for cell in [ws[f'A{row}'], ws[f'B{row}'], ws[f'C{row}']]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        row += 1
        
        # Data
        for cmd in unhandled:
            # Count files containing this command
            files_with_cmd = [fname for fname, cmds in self.file_commands.items() 
                            if cmd in cmds]
            
            ws[f'A{row}'] = cmd
            ws[f'B{row}'] = self.command_counts[cmd]
            ws[f'C{row}'] = len(files_with_cmd)
            
            # Highlight
            for col in range(1, 4):
                ws.cell(row=row, column=col).fill = PatternFill(
                    start_color="FFE6E6", end_color="FFE6E6", fill_type="solid"
                )
            
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 30


def main():
    """Main entry point"""
    # Setup paths
    script_dir = Path(__file__).parent
    batch_tests_dir = script_dir / "batch_tests"
    output_file = script_dir / "command_coverage_report.xlsx"
    
    # Verify batch_tests directory exists
    if not batch_tests_dir.exists():
        print(f"Error: Directory not found: {batch_tests_dir}")
        sys.exit(1)
    
    print("=" * 80)
    print("CGM Command Coverage Analyzer")
    print("=" * 80)
    print(f"Analyzing files in: {batch_tests_dir}")
    print()
    
    # Create analyzer
    analyzer = CommandAnalyzer(batch_tests_dir)
    
    # Analyze all files
    analyzer.analyze_all_files()
    
    # Generate report
    analyzer.generate_excel_report(output_file)
    
    print()
    print("=" * 80)
    print("Analysis Complete")
    print("=" * 80)
    
    # Summary
    total_commands = len(analyzer.all_commands)
    handled_commands = sum(1 for cmd in analyzer.all_commands 
                          if analyzer.get_command_mapping(cmd) != "NOT HANDLED")
    
    print(f"Total unique commands: {total_commands}")
    print(f"Handled commands: {handled_commands}")
    print(f"Unhandled commands: {total_commands - handled_commands}")
    print(f"Coverage: {(handled_commands/total_commands*100):.1f}%")
    print()
    print(f"Report saved to: {output_file}")


if __name__ == "__main__":
    main()
